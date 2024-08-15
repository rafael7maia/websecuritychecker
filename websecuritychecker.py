import requests
import ssl
import socket
import argparse
import urllib3
import re
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Desativar avisos de SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Lista de cabeçalhos de segurança a serem verificados
SECURITY_HEADERS = [
    "Strict-Transport-Security",
    "X-Frame-Options",
    "X-XSS-Protection",
    "X-Content-Type-Options",
    "Referrer-Policy",
    "Content-Security-Policy",
    "Permissions-Policy",
    "Cache-Control"
]

# Função para garantir que a URL tenha um esquema
def ensure_url_scheme(url):
    if not url.startswith(('http://', 'https://')):
        url = 'http://' + url  # Adiciona http:// se nenhum esquema estiver presente
    return url

# Verificação de Resolução de Nome de Domínio
def is_domain_resolvable(url):
    try:
        hostname = url.split("//")[-1].split("/")[0]
        socket.gethostbyname(hostname)
        return True
    except socket.error:
        return False

# Verificação de Cabeçalhos de Segurança
def check_security_headers(headers):
    results = []
    for header in SECURITY_HEADERS:
        if header in headers:
            results.append(f"[OK] {header}: {headers[header]}")
        else:
            if header == "Content-Security-Policy":
                results.append(f"[Vulnerável] Content Security Policy não encontrado")
            else:
                results.append(f"[Vulnerável] {header}: não encontrado")
    return results

# Verificação de TLS/SSL
def check_tls(url):
    hostname = url.split("//")[-1].split("/")[0]
    context = ssl.create_default_context()
    with socket.create_connection((hostname, 443)) as sock:
        with context.wrap_socket(sock, server_hostname=hostname) as secure_sock:
            cipher = secure_sock.cipher()
            tls_version = secure_sock.version()  # Captura a versão do TLS
    sweet32_vulnerable = "3DES" in cipher[0] or "DES" in cipher[0]
    cbc_cipher = "CBC" in cipher[0]
    beast_vulnerable = tls_version == "TLSv1.0"
    deprecated_versions = check_tls_versions([tls_version])

    return [cipher[0]], sweet32_vulnerable, cbc_cipher, beast_vulnerable, deprecated_versions

# Verificação de Versões TLS
def check_tls_versions(supported_versions):
    deprecated_versions = []
    if "TLSv1.0" in supported_versions:
        deprecated_versions.append("TLSv1.0")
    if "TLSv1.1" in supported_versions:
        deprecated_versions.append("TLSv1.1")
    return deprecated_versions

# Verificação de Cookies
def check_cookies(cookies):
    insecure_cookies = []
    inconsistent_cookies = []
    for cookie in cookies:
        if not cookie.secure:
            insecure_cookies.append(cookie.name)
        if cookie.secure and not cookie.has_nonstandard_attr('HttpOnly'):
            inconsistent_cookies.append(cookie.name)
    return insecure_cookies, inconsistent_cookies

# Verificação de Bibliotecas JavaScript Vulneráveis
def check_vulnerable_js(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    scripts = soup.find_all('script', src=True)
    vulnerable_libraries = []
    for script in scripts:
        src = script['src']
        if 'jquery' in src.lower() and 'jquery-1.' in src.lower():
            vulnerable_libraries.append(f"jQuery (versão antiga): {src}")
    return vulnerable_libraries

# Verificação de ViewState
def check_viewstate(html_content):
    return "__VIEWSTATE" in html_content

# Verificação de Mensagens de Erro
def check_error_messages(html_content):
    error_keywords = ["exception", "error occurred", "stack trace", "runtime error"]
    return any(keyword in html_content.lower() for keyword in error_keywords)

# Verificação de Cabeçalhos de Versão
def check_version_headers(headers):
    version_headers = ["Server", "X-Powered-By", "X-AspNet-Version", "X-AspNetMvc-Version"]
    results = []
    for header in version_headers:
        if header in headers:
            if headers[header].strip():  # Verifica se o cabeçalho não está vazio
                results.append(f"[Vulnerável] {header}: {headers[header]}")
            else:
                results.append(f"[OK] {header}: encontrado, mas vazio")
        else:
            results.append(f"[OK] {header}: não encontrado")
    return results

# Verificação de Proteção contra Clickjacking
def check_clickjacking(headers):
    return "X-Frame-Options" not in headers

# Verificação de Divulgação de IP Interno
def check_internal_ip_disclosure(html_content):
    ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    return bool(re.search(ip_pattern, html_content))

# Verificação de Divulgação de Erro ASP.NET
def check_aspnet_error_disclosure(html_content):
    return "ASP.NET Error" in html_content

# Verificação de Vulnerabilidades a Injeções CGI
def check_cgi_injections(url):
    test_url = f"{url}?test=<script>alert('XSS')</script>"
    response = requests.get(test_url, verify=False)
    return "<script>alert('XSS')</script>" in response.text

# Verificação de Divulgação Detalhada de Erros
def check_detailed_error_disclosure(html_content):
    error_keywords = ["exception details", "stack trace", "line number", "source file"]
    return any(keyword in html_content.lower() for keyword in error_keywords)

# Geração do Relatório em Excel
def generate_excel_report(results, url):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Segurança"

    # Estilos
    title_font = Font(size=14, bold=True)
    header_font = Font(size=12, bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Título
    ws['A1'] = f"Relatório de Segurança para {url}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')  # Ajustado para 4 colunas

    # Cabeçalhos
    headers = ["Verificação", "Resultado", "Recomendação", "Como Implementar"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 4
    for section, section_results in results.items():
        ws.cell(row=row, column=1, value=section)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Ordenar resultados: vulneráveis primeiro, depois OK
        sorted_section_results = sorted(section_results, key=lambda x: (x.startswith("[OK]"), x))

        for result in sorted_section_results:
            if result.startswith("[Vulnerável]"):
                status = "Vulnerável"
                recommendation, how_to = get_recommendation(result)
            elif result.startswith("[OK]"):
                status = "OK"
                recommendation = "Nenhuma ação necessária."
                how_to = ""
            else:
                status = "Informação"
                recommendation = "Revisar e tomar ação se necessário."
                how_to = ""

            ws.cell(row=row, column=1, value=result.split(":")[0].replace("[Vulnerável]", "").replace("[OK]", "").strip())
            ws.cell(row=row, column=2, value=status)
            ws.cell(row=row, column=3, value=recommendation)
            ws.cell(row=row, column=4, value=how_to)  # Adicionando a coluna "Como Implementar"

            if status == "Vulnerável":
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            row += 1

    # Ajustar largura das colunas
    for col in range(1, 5):
        max_length = 0
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=4, max_row=row):
            for c in cell:
                if c.value:
                    max_length = max(max_length, len(str(c.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width

    filename = f"security_report_{url.replace('https://', '').replace('http://', '').replace('/', '_')}.xlsx"
    wb.save(filename)
    print(f"\nRelatório gerado: {filename}")

# Função para obter recomendações com base nos resultados
def get_recommendation(result):
    if "Content Security Policy" in result:
        return (
            "Implemente uma política de segurança de conteúdo.",
            "Exemplo de implementação:\nAdicione o seguinte cabeçalho HTTP ao seu servidor:\nContent-Security-Policy: default-src 'self'; img-src 'self' data:; script-src 'self';"
        )
    elif "Strict-Transport-Security" in result:
        return ("Configure o cabeçalho HSTS para forçar conexões HTTPS.",
                "Adicione o cabeçalho 'Strict-Transport-Security' nas respostas HTTP. Exemplo: "
                "Strict-Transport-Security: max-age=31536000; includeSubDomains; preload")
    elif "X-Frame-Options" in result:
        return ("Configure o cabeçalho X-Frame-Options para prevenir clickjacking.",
                "Adicione o cabeçalho 'X-Frame-Options' nas respostas HTTP. Exemplo: "
                "X-Frame-Options: DENY ou X-Frame-Options: SAMEORIGIN")
    elif "X-XSS-Protection" in result:
        return ("Ative a proteção XSS do navegador.",
                "Adicione o cabeçalho 'X-XSS-Protection' nas respostas HTTP. Exemplo: "
                "X-XSS-Protection: 1; mode=block")
    elif "X-Content-Type-Options" in result:
        return ("Configure o cabeçalho X-Content-Type-Options para 'nosniff'.",
                "Adicione o cabeçalho 'X-Content-Type-Options' nas respostas HTTP. Exemplo: "
                "X-Content-Type-Options: nosniff")
    elif "Referrer-Policy" in result:
        return ("Configure uma política de referência adequada.",
                "Adicione o cabeçalho 'Referrer-Policy' nas respostas HTTP. Exemplo: "
                "Referrer-Policy: no-referrer-when-downgrade")
    elif "Permissions-Policy" in result:
        return ("Implemente uma política de permissões para controlar recursos do navegador.",
                "Adicione o cabeçalho 'Permissions-Policy' nas respostas HTTP. Exemplo: "
                "Permissions-Policy: geolocation=(self), microphone=()")
    elif "Cache-Control" in result:
        return ("Configure o cabeçalho Cache-Control para controlar o cache do navegador.",
                "Adicione o cabeçalho 'Cache-Control' nas respostas HTTP. Exemplo: "
                "Cache-Control: no-store, no-cache, must-revalidate")
    elif "Sweet32" in result:
        return ("Desative cifras DES e 3DES vulneráveis à Sweet32.",
                "Configure o servidor para não usar cifras DES e 3DES. Exemplo: "
                "No Apache: SSLCipherSuite HIGH:!3DES:!DES")
    elif "Cookies inseguros" in result:
        return ("Configure todos os cookies sensíveis com o atributo 'Secure'.",
                "Adicione o atributo 'Secure' aos cookies sensíveis. Exemplo: "
                "Set-Cookie: sessionid=abc123; Secure; HttpOnly")
    elif "Cookies inconsistentes" in result:
        return ("Adicione o atributo 'HttpOnly' a todos os cookies seguros.",
                "Adicione o atributo 'HttpOnly' aos cookies. Exemplo: "
                "Set-Cookie: sessionid=abc123; HttpOnly")
    elif "jQuery" in result:
        return ("Atualize para a versão mais recente do jQuery.",
                "Substitua a versão antiga do jQuery pela versão mais recente em seu código. Exemplo: "
                "<script src='https://code.jquery.com/jquery-3.6.0.min.js'></script>")
    elif "ViewState" in result:
        return ("Certifique-se de que o ViewState esteja criptografado e use validação MAC.",
                "Ative a criptografia do ViewState no ASP.NET. Exemplo: "
                "EnableViewStateMac = true; ViewStateEncryptionMode = ViewStateEncryptionMode.Enabled;")
    elif "Mensagens de erro detalhadas" in result:
        return ("Desative mensagens de erro detalhadas em produção.",
                "Configure o ambiente para não exibir detalhes de erro. Exemplo: "
                "No ASP.NET: customErrors mode='On'")
    elif "Server:" in result or "X-Powered-By:" in result:
        return ("Remova ou oculte cabeçalhos que revelam informações de versão.",
                "Configure o servidor para não enviar cabeçalhos de versão. Exemplo: "
                "No Apache: ServerTokens Prod; ServerSignature Off")
    elif "Proteção contra Clickjacking" in result:
        return ("Implemente proteção contra clickjacking usando X-Frame-Options ou CSP.",
                "Adicione o cabeçalho 'X-Frame-Options' ou 'Content-Security-Policy' nas respostas HTTP.")
    elif "Divulgação de IP interno" in result:
        return ("Evite expor endereços IP internos no conteúdo da página.",
                "Revise o código para garantir que informações sensíveis não sejam expostas.")
    elif "Divulgação de erro ASP.NET" in result:
        return ("Desative a exibição de erros detalhados do ASP.NET em produção.",
                "Configure o ambiente para não exibir detalhes de erro. Exemplo: "
                "No ASP.NET: customErrors mode='On'")
    elif "Vulnerabilidade a injeções CGI" in result:
        return ("Implemente validação e sanitização adequadas para parâmetros de entrada.",
                "Revise o código para garantir que os parâmetros de entrada sejam validados e sanitizados.")
    elif "Divulgação detalhada de erros" in result:
        return ("Configure o aplicativo para não exibir detalhes de erro em produção.",
                "Configure o ambiente para não exibir detalhes de erro. Exemplo: "
                "No ASP.NET: customErrors mode='On'")
    else:
        return ("Revise e corrija a vulnerabilidade identificada.", "Considere consultar a documentação de segurança relevante.")

# Função principal para verificar a segurança da URL
def check_security(url):
    url = ensure_url_scheme(url)  # Garantir que a URL tenha esquema
    if not is_domain_resolvable(url):
        print(f"Erro: Não é possível resolver o domínio para {url}. Verifique se o domínio está correto.")
        return
    
    try:
        response = requests.get(url, verify=False, timeout=10)  # Definindo um tempo limite de 10 segundos
        headers = response.headers
        print(f"Verificando cabeçalhos de segurança e vulnerabilidades para: {url}\n")

        results = {
            "Cabeçalhos de Segurança": check_security_headers(headers),
            "Cabeçalhos de Versão": check_version_headers(headers),
            "TLS/SSL Verificações": [],
            "Cookies": [],
            "Bibliotecas JavaScript Vulneráveis": [],
            "Outras Verificações": []
        }

        # Verificações TLS/SSL
        cipher_suites, sweet32_vulnerable, cbc_cipher, beast_vulnerable, deprecated_versions = check_tls(url)
        results["TLS/SSL Verificações"].append(f"[OK] Cipher: {cipher_suites[0]}")
        results["TLS/SSL Verificações"].append(f"[{'Vulnerável' if sweet32_vulnerable else 'OK'}] Sweet32 Vulnerável: {'Sim' if sweet32_vulnerable else 'Não'}")
        results["TLS/SSL Verificações"].append(f"[{'Vulnerável' if cbc_cipher else 'OK'}] Cipher CBC: {'Usado' if cbc_cipher else 'Não usado'}")
        results["TLS/SSL Verificações"].append(f"[{'Vulnerável' if beast_vulnerable else 'OK'}] Vulnerabilidade BEAST: {'Sim' if beast_vulnerable else 'Não'}")

        if deprecated_versions:
            results["TLS/SSL Verificações"].append(f"[Vulnerável] Versões TLS depreciadas encontradas: {', '.join(deprecated_versions)}")

        # Verificações de cookies
        insecure_cookies, inconsistent_cookies = check_cookies(response.cookies)
        if insecure_cookies:
            results["Cookies"].append(f"[Vulnerável] Cookies inseguros: {', '.join(insecure_cookies)}")
        if inconsistent_cookies:
            results["Cookies"].append(f"[Vulnerável] Cookies inconsistentes: {', '.join(inconsistent_cookies)}")
        if not insecure_cookies and not inconsistent_cookies:
            results["Cookies"].append("[OK] Nenhum problema encontrado com cookies")

        # Verificação de bibliotecas JavaScript vulneráveis
        vulnerable_libraries = check_vulnerable_js(response.text)
        if vulnerable_libraries:
            results["Bibliotecas JavaScript Vulneráveis"].extend([f"[Vulnerável] Biblioteca vulnerável encontrada: {lib}" for lib in vulnerable_libraries])
        else:
            results["Bibliotecas JavaScript Vulneráveis"].append("[OK] Nenhuma biblioteca JavaScript vulnerável conhecida encontrada")

        # Outras verificações
        results["Outras Verificações"].append(f"[{'Vulnerável' if check_viewstate(response.text) else 'OK'}] ViewState: {'Encontrado' if check_viewstate(response.text) else 'Não encontrado'}")
        results["Outras Verificações"].append(f"[{'Vulnerável' if check_error_messages(response.text) else 'OK'}] Mensagens de erro detalhadas: {'Encontradas' if check_error_messages(response.text) else 'Não encontradas'}")
        results["Outras Verificações"].append(f"[{'Vulnerável' if check_clickjacking(headers) else 'OK'}] Proteção contra Clickjacking: {'Ausente' if check_clickjacking(headers) else 'Presente'}")
        results["Outras Verificações"].append(f"[{'Vulnerável' if check_internal_ip_disclosure(response.text) else 'OK'}] Divulgação de IP interno: {'Encontrada' if check_internal_ip_disclosure(response.text) else 'Não encontrada'}")
        results["Outras Verificações"].append(f"[{'Vulnerável' if check_aspnet_error_disclosure(response.text) else 'OK'}] Divulgação de erro ASP.NET: {'Encontrada' if check_aspnet_error_disclosure(response.text) else 'Não encontrada'}")
        results["Outras Verificações"].append(f"[{'Vulnerável' if check_cgi_injections(url) else 'OK'}] Vulnerabilidade a injeções CGI: {'Presente' if check_cgi_injections(url) else 'Ausente'}")
        results["Outras Verificações"].append(f"[{'Vulnerável' if check_detailed_error_disclosure(response.text) else 'OK'}] Divulgação detalhada de erros: {'Encontrada' if check_detailed_error_disclosure(response.text) else 'Não encontrada'}")

        # Imprimir resultados na tela
        print(f"\nResultados para {url}:\n")
        for section, section_results in results.items():
            print(f"{section}:")
            # Ordenar resultados: vulneráveis primeiro, depois OK
            sorted_section_results = sorted(section_results, key=lambda x: (x.startswith("[OK]"), x))

            for result in sorted_section_results:
                print(f"  - {result}")
            print()  # Linha em branco entre seções

        # Perguntar se o usuário deseja gerar o relatório em Excel
        generate_excel = input("Deseja gerar um relatório em Excel? (s/n): ").strip().lower()
        if generate_excel == 's':
            generate_excel_report(results, url)

    except requests.exceptions.RequestException as e:
        print(f"Erro ao acessar {url}: {e}")

# Função para ler URLs de um arquivo
def read_urls_from_file(file_path):
    try:
        with open(file_path, 'r') as file:
            urls = file.readlines()
            # Remover espaços em branco e quebras de linha
            urls = [url.strip() for url in urls]
            # Remover URLs duplicadas
            unique_urls = list(set(urls))
            print(f"O arquivo possui {len(unique_urls)} sites únicos.")
            return unique_urls
    except FileNotFoundError:
        print(f"Erro: O arquivo {file_path} não foi encontrado.")
        return []

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Verificador de segurança de cabeçalhos e TLS')
    parser.add_argument('url', nargs='?', help='URL a ser verificada')
    parser.add_argument('-l', '--list', help='Caminho para o arquivo de lista de URLs')
    args = parser.parse_args()

    if args.list:
        urls = read_urls_from_file(args.list)
        for url in urls:
            check_security(url)
    elif args.url:
        check_security(args.url)
    else:
        print("Por favor, forneça uma URL ou um arquivo de lista de URLs para verificar.")

